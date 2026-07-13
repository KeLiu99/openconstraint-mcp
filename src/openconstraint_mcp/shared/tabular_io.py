"""Bounded, local tabular I/O for ``.xlsx`` and ``.csv`` files.

Mechanical I/O only: this leaf moves scalars between a file and a
``TabularData``/``TabularWriteResult`` model. It never infers what a column
*means*, never evaluates a formula, and never calls out to the network or a
subprocess. Interpreting columns is the client's job.

Dependencies: stdlib + openpyxl (lazily, only where actually used) +
``schemas.tabular`` + ``shared.hashing`` + ``shared.path_checks``.
Deliberately does not import ``shared.save_target``: a single output file has
no manifest and no managed-directory policy, only the low-level "absolute,
right kind, parent exists" shape that ``shared.path_checks`` factors out.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
import math
import os
import re
import sys
import tempfile
from collections.abc import Callable, Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any

from ..schemas.tabular import TabularCell, TabularData, TabularFormat, TabularWriteResult
from .hashing import path_sha256
from .path_checks import resolve_absolute_target

# The hard ceiling on a serialized TabularData body, independent of max_rows.
# Only whole rows are ever returned, so a page is trimmed to fit rather than
# a cell being cut in half.
MAX_TABULAR_RESPONSE_BYTES: int = 1_048_576

DEFAULT_MAX_ROWS: int = 1000

# Excel's hard per-cell string limit. openpyxl silently TRUNCATES a longer
# string (see Cell.check_string), which would be a silent data change — so we
# reject the string up front instead of letting it through.
XLSX_MAX_STRING_LENGTH: int = 32_767

# A leading one of these makes a spreadsheet treat text as a formula. XLSX can
# store the text as an explicit string cell, so it is written verbatim there.
# CSV has no way to encode "this is literal text", so such strings are refused.
CSV_FORMULA_PREFIXES: tuple[str, ...] = ("=", "+", "-", "@")

WRITE_SHEET_NAME: str = "Sheet1"

_SUFFIX_FORMATS: dict[str, TabularFormat] = {".xlsx": "xlsx", ".csv": "csv"}

# Python's csv module caps one field at 128 KiB and raises an opaque _csv.Error
# past it — which is not even a ValueError, so it would escape the tools' error
# translation. The response ceiling is the limit that should govern a read, so
# raise csv's cap to match it. Done once at import rather than per read: reads
# run in a thread pool, where a per-call set/restore would race (one call's
# restore would drop the cap under another call still parsing). This is an
# idempotent bump of a stdlib knob, not app state — a field past even this is
# still translated to a bounded ValueError below.
if csv.field_size_limit() < MAX_TABULAR_RESPONSE_BYTES:
    csv.field_size_limit(MAX_TABULAR_RESPONSE_BYTES)


def _format_for(path: Path) -> TabularFormat:
    """Return the tabular format for ``path``'s suffix, or raise ``ValueError``."""
    fmt = _SUFFIX_FORMATS.get(path.suffix.lower())
    if fmt is None:
        accepted = ", ".join(sorted(_SUFFIX_FORMATS))
        raise ValueError(
            f"unsupported tabular file type {path.suffix!r} for {path}; expected one of {accepted}"
        )
    return fmt


def validate_tabular_read_path(path: Path) -> Path:
    """Resolve ``path`` and require an existing ``.xlsx``/``.csv`` regular file."""
    resolved = path.expanduser().resolve()
    _format_for(resolved)
    if not resolved.exists():
        raise ValueError(f"tabular file does not exist: {resolved}")
    if not resolved.is_file():
        raise ValueError(f"tabular path is not a regular file: {resolved}")
    return resolved


def validate_tabular_write_path(path: Path) -> Path:
    """Resolve ``path`` and require an absolute ``.xlsx``/``.csv`` target in an existing dir."""
    resolved = resolve_absolute_target(
        path, arg_name="target_path", kind="regular file", is_valid_kind=Path.is_file
    )
    _format_for(resolved)
    return resolved


# --- reading -------------------------------------------------------------------


def _normalize_header(value: object, index: int) -> str:
    """Return the string header for column ``index`` (zero-based).

    Headers are always strings. A blank position — ``None`` or an empty string
    — becomes the positional name ``col_<n>`` (one-based). Dates/times render
    ISO-8601; any other non-string renders through ``str``. Duplicate non-blank
    names are preserved: deduplicating them would be interpretation, and the
    client may legitimately have two columns of the same name.
    """
    if value is None:
        return f"col_{index + 1}"
    if isinstance(value, str):
        return value if value != "" else f"col_{index + 1}"
    if isinstance(value, dt.date | dt.time):
        return value.isoformat()
    return str(value)


def _normalize_cell(value: object) -> TabularCell:
    """Coerce one raw cell value to a JSON scalar.

    Strings, booleans, integers, and finite floats pass through with their type
    intact. Dates/times become ISO-8601 strings. Anything else — a
    ``timedelta`` from an elapsed-time cell, a stray non-finite float — renders
    through ``str`` rather than failing the read, since a value that cannot be
    a JSON scalar would otherwise make the whole page unreadable.
    """
    if value is None or isinstance(value, str | bool | int):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if isinstance(value, dt.date | dt.time):
        return value.isoformat()
    return str(value)


def _csv_records(reader: Any, path: Path) -> Iterator[list[Any]]:
    """Yield the reader's records, translating a parse failure to a ``ValueError``.

    ``_csv.Error`` is not a ``ValueError``, so it would bypass the tools' error
    translation and surface as an opaque crash. The realistic trigger is a field
    past the size cap raised above — i.e. a field too large to return anyway.
    ``next(reader)`` also pulls more bytes from the underlying file, so a
    removable/network filesystem failing mid-read raises a raw ``OSError`` here
    too — translate that the same way.
    """
    while True:
        try:
            record = next(reader)
        except StopIteration:
            return
        except csv.Error as exc:
            raise ValueError(
                f"cannot parse {path} at line {reader.line_num}: {exc}. A single CSV field "
                f"must stay under the {MAX_TABULAR_RESPONSE_BYTES}-byte response limit."
            ) from exc
        except OSError as exc:
            raise ValueError(f"cannot read {path}: {exc}") from exc
        yield record


@contextmanager
def _open_csv_rows(path: Path) -> Iterator[Iterator[list[Any]]]:
    """Yield an iterator of CSV records, using one fixed dialect.

    Comma-separated, ``"``-quoted, UTF-8 — no dialect sniffing. ``utf-8-sig``
    strips a byte-order mark if present (spreadsheet apps routinely write one),
    which is an encoding concern, not a dialect one; a plain UTF-8 file decodes
    identically.

    ``validate_tabular_read_path`` already confirms the file exists, but a
    permission-denied (or otherwise unopenable) file still raises ``OSError``
    from ``open()`` itself — not a ``ValueError`` — so it would otherwise
    bypass the tools' error translation.
    """
    try:
        handle = path.open("r", encoding="utf-8-sig", newline="")
    except OSError as exc:
        raise ValueError(f"cannot read {path}: {exc}") from exc
    try:
        yield _csv_records(csv.reader(handle), path)
    finally:
        # Guarded like the write path's staging cleanup: a close() failure
        # (e.g. a stale network-filesystem handle) must not replace whatever
        # the body above already raised, per Python's finally-exception
        # -replaces-try-exception semantics — only surface it when nothing
        # else is already propagating. The check must run BEFORE the close()
        # call: inside the `except OSError` below, sys.exc_info() always
        # reports that very close() failure as "currently being handled", so
        # checking it there would be a tautology that never fires.
        already_failing = sys.exc_info()[0] is not None
        try:
            handle.close()
        except OSError as exc:
            if not already_failing:
                raise ValueError(f"cannot read {path}: {exc}") from exc


def _xlsx_records(worksheet: Any, path: Path) -> Iterator[list[Any]]:
    """Yield ``worksheet``'s rows, translating a parse failure to a ``ValueError``.

    ``read_only=True`` streams the sheet rather than parsing it up front, so a
    corrupt worksheet's XML raises only once its rows are actually iterated —
    here, not from the ``load_workbook`` call in ``_open_xlsx_rows``. The
    raised type (``xml.etree.ElementTree.ParseError`` and others) is not a
    ``ValueError``, so it would otherwise bypass the tools' error translation.
    """
    try:
        for row in worksheet.iter_rows(values_only=True):
            yield list(row)
    except Exception as exc:
        raise ValueError(f"cannot read {path} as an XLSX workbook: {exc}") from exc


@contextmanager
def _open_xlsx_rows(
    path: Path, sheet: str | None
) -> Iterator[tuple[str, list[str], Iterator[list[Any]]]]:
    """Yield ``(sheet_name, available_sheets, rows)`` for one worksheet.

    ``data_only=True`` reads a formula cell's CACHED result — the server never
    evaluates a formula, so a formula that was never calculated reads as
    ``None``. ``read_only=True`` streams the sheet rather than building the
    whole object graph, so the workbook must stay open while rows are consumed
    — hence the context manager.

    A malformed file (not a zip at all, or a zip missing the parts an XLSX
    workbook requires) raises an assortment of exception types from openpyxl's
    zip/XML layers — ``zipfile.BadZipFile``, ``KeyError``, and others — none of
    which is a ``ValueError``, so none would reach the tools' error
    translation. Translate every failure from this one call uniformly instead.

    ``workbook.sheetnames`` covers every sheet, including chart sheets — which
    have no rows and no ``iter_rows``. ``available`` is restricted to
    ``workbook.worksheets`` (data sheets only) so a chart sheet is never
    offered as a selectable name that would only fail later. The default
    selection is ``workbook.active`` when that is itself a data worksheet —
    preserving the documented "defaults to the active sheet" behavior — and
    falls back to the first data worksheet only when the active sheet is a
    chart sheet.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"cannot read {path} as an XLSX workbook: {exc}") from exc
    try:
        data_sheets = workbook.worksheets
        available = [str(ws.title) for ws in data_sheets]
        if sheet is None:
            if not data_sheets:
                if workbook.sheetnames:
                    raise ValueError(f"{path} has no worksheets (only chart sheets)")
                raise ValueError(f"{path} has no worksheets")
            active = workbook.active
            worksheet = active if active in data_sheets else data_sheets[0]
        else:
            if sheet not in available:
                if sheet in workbook.sheetnames:
                    raise ValueError(
                        f"sheet {sheet!r} in {path} is a chart sheet, not a data "
                        f"worksheet; available data sheets: {', '.join(available)}"
                    )
                names = ", ".join(available)
                raise ValueError(f"sheet {sheet!r} not found in {path}; available sheets: {names}")
            worksheet = workbook[sheet]
        yield str(worksheet.title), available, _xlsx_records(worksheet, path)
    finally:
        workbook.close()


def _row_estimate(row: list[TabularCell]) -> int:
    """Approximate the serialized byte cost of one row, plus its separating comma.

    A cheap streaming bound so a huge ``max_rows`` cannot buffer an unbounded
    page into memory. It ignores the body's fixed overhead (headers, metadata),
    which makes it a slight UNDER-estimate of the true cost — deliberate, so it
    can never stop collecting a row that would in fact have fit. The exact
    ceiling is enforced afterward against the real serialized body.
    """
    return len(json.dumps(row, ensure_ascii=False, separators=(",", ":")).encode("utf-8")) + 1


def _build_page(
    *,
    headers: list[str],
    rows: list[list[TabularCell]],
    sheet_name: str | None,
    available_sheets: list[str],
    row_offset: int,
    total_rows: int,
    byte_limited: bool,
) -> TabularData:
    """Assemble a ``TabularData`` whose pagination metadata matches its rows."""
    consumed = row_offset + len(rows)
    if consumed >= total_rows:
        return TabularData(
            headers=headers,
            rows=rows,
            sheet_name=sheet_name,
            available_sheets=available_sheets,
            row_offset=row_offset,
            next_row_offset=None,
            total_rows=total_rows,
            truncated=False,
            truncation_reason=None,
        )
    return TabularData(
        headers=headers,
        rows=rows,
        sheet_name=sheet_name,
        available_sheets=available_sheets,
        row_offset=row_offset,
        next_row_offset=consumed,
        total_rows=total_rows,
        truncated=True,
        truncation_reason="max_bytes" if byte_limited else "max_rows",
    )


def _body_size(page: TabularData) -> int:
    """Return the exact UTF-8 byte length of the page's serialized JSON body."""
    return len(page.model_dump_json().encode("utf-8"))


class _Scan:
    """What one streaming pass over the rows collected.

    ``header_values`` is the raw header record (``None`` when ``has_header`` is
    false or the file is empty), ``page`` the buffered rows for this request,
    ``total_rows`` every data row in the file, ``width`` the widest data row
    (used only to name a headerless file's columns), and ``byte_limited``
    whether buffering stopped on the byte estimate rather than ``max_rows``.
    """

    def __init__(self) -> None:
        self.header_values: list[Any] | None = None
        self.page: list[list[TabularCell]] = []
        self.total_rows: int = 0
        self.width: int = 0
        self.byte_limited: bool = False


def _scan_rows(
    raw_rows: Iterator[list[Any]],
    *,
    has_header: bool,
    row_offset: int,
    max_rows: int,
) -> _Scan:
    """Stream every row once, buffering only the requested page.

    The full pass is unavoidable — ``total_rows`` and reaching ``row_offset``
    both require it — but memory stays bounded by the page, not the file: rows
    before the offset, after ``max_rows``, and past the byte estimate are
    counted and discarded.
    """
    scan = _Scan()
    buffering = True
    estimate = 0
    for index, raw_row in enumerate(raw_rows):
        if index == 0 and has_header:
            scan.header_values = list(raw_row)
            continue
        data_index = scan.total_rows
        scan.total_rows += 1
        scan.width = max(scan.width, len(raw_row))
        if not buffering or data_index < row_offset or len(scan.page) >= max_rows:
            continue
        row = [_normalize_cell(value) for value in raw_row]
        scan.page.append(row)
        estimate += _row_estimate(row)
        if estimate > MAX_TABULAR_RESPONSE_BYTES:
            # This row already busts the ceiling, so no later row can fit
            # either. Stop buffering (but keep counting); the exact trim
            # decides whether even this row survives.
            scan.byte_limited = True
            buffering = False
    return scan


def _scan_source(
    resolved: Path,
    sheet: str | None,
    *,
    has_header: bool,
    row_offset: int,
    max_rows: int,
) -> tuple[_Scan, str | None, list[str]]:
    """Scan ``resolved`` and return ``(scan, sheet_name, available_sheets)``.

    A CSV has no sheets, so it always returns ``(scan, None, [])``.
    """
    if _format_for(resolved) == "csv":
        if sheet is not None:
            raise ValueError(f"a CSV file has no sheets, so sheet={sheet!r} cannot be selected")
        with _open_csv_rows(resolved) as csv_rows:
            scan = _scan_rows(
                csv_rows, has_header=has_header, row_offset=row_offset, max_rows=max_rows
            )
        return scan, None, []
    with _open_xlsx_rows(resolved, sheet) as (sheet_name, available_sheets, xlsx_rows):
        scan = _scan_rows(
            xlsx_rows, has_header=has_header, row_offset=row_offset, max_rows=max_rows
        )
    return scan, sheet_name, available_sheets


def read_tabular_data(
    path: Path,
    *,
    sheet: str | None = None,
    has_header: bool = True,
    row_offset: int = 0,
    max_rows: int = DEFAULT_MAX_ROWS,
) -> TabularData:
    """Read one bounded page of rows from a ``.xlsx`` or ``.csv`` file.

    ``row_offset`` is a zero-based offset among data rows (the header, when
    present, is not a data row); ``max_rows`` caps the page. The serialized
    body is additionally capped at ``MAX_TABULAR_RESPONSE_BYTES`` — whichever
    bound first is reported as ``truncation_reason``, and the page always
    contains whole, unmodified rows. Every later page is directly reachable via
    the returned ``next_row_offset``.

    Scan cost is proportional to the whole file, not the page: reaching an
    offset and counting ``total_rows`` both require streaming from the start.
    """
    if row_offset < 0:
        raise ValueError(f"row_offset must be >= 0 (got {row_offset})")
    if max_rows < 1:
        raise ValueError(f"max_rows must be >= 1 (got {max_rows})")

    resolved = validate_tabular_read_path(path)
    scan, sheet_name, available_sheets = _scan_source(
        resolved, sheet, has_header=has_header, row_offset=row_offset, max_rows=max_rows
    )

    # Headers must be identical on every page, so a headerless file derives its
    # positional names from the WIDEST row in the file — not from whichever
    # rows this page happens to contain. Likewise, a header row narrower than
    # some later data row (ragged CSV) still needs a name for every cell any
    # page could contain, so the header count covers the wider of the two.
    if scan.header_values is not None:
        header_count = max(len(scan.header_values), scan.width)
        headers = [
            _normalize_header(
                scan.header_values[index] if index < len(scan.header_values) else None, index
            )
            for index in range(header_count)
        ]
    else:
        headers = [f"col_{index + 1}" for index in range(scan.width)]

    candidate = _build_page(
        headers=headers,
        rows=scan.page,
        sheet_name=sheet_name,
        available_sheets=available_sheets,
        row_offset=row_offset,
        total_rows=scan.total_rows,
        byte_limited=scan.byte_limited,
    )
    if _body_size(candidate) <= MAX_TABULAR_RESPONSE_BYTES:
        return candidate
    return _trim_to_byte_ceiling(
        headers=headers,
        rows=scan.page,
        sheet_name=sheet_name,
        available_sheets=available_sheets,
        row_offset=row_offset,
        total_rows=scan.total_rows,
    )


def _trim_to_byte_ceiling(
    *,
    headers: list[str],
    rows: list[list[TabularCell]],
    sheet_name: str | None,
    available_sheets: list[str],
    row_offset: int,
    total_rows: int,
) -> TabularData:
    """Drop trailing whole rows until the serialized body fits the ceiling.

    Only called once the full candidate is known to be over the limit, so every
    page considered here is a truncated one — a fixed metadata shape, under
    which body size is monotone in row count. That makes a binary search exact
    and bounded to ~log2(len(rows)) serializations, instead of the quadratic
    re-serialize-per-popped-row a linear scan would cost on a large page.

    Refuses rather than returning a page that makes no forward progress: a
    caller that received zero rows and the same offset back would loop forever.
    """

    def page_of(count: int) -> TabularData:
        return _build_page(
            headers=headers,
            rows=rows[:count],
            sheet_name=sheet_name,
            available_sheets=available_sheets,
            row_offset=row_offset,
            total_rows=total_rows,
            byte_limited=True,
        )

    if _body_size(page_of(0)) > MAX_TABULAR_RESPONSE_BYTES:
        raise ValueError(
            f"the normalized headers alone exceed the {MAX_TABULAR_RESPONSE_BYTES}-byte "
            f"response limit, so no page of this file can be returned"
        )

    low, high = 0, len(rows) - 1  # page_of(0) fits; the full page does not.
    while low < high:
        middle = (low + high + 1) // 2
        if _body_size(page_of(middle)) <= MAX_TABULAR_RESPONSE_BYTES:
            low = middle
        else:
            high = middle - 1
    if low == 0:
        raise ValueError(
            f"the data row at offset {row_offset} does not fit the "
            f"{MAX_TABULAR_RESPONSE_BYTES}-byte response limit on its own, so no page "
            f"starting there can make progress"
        )
    return page_of(low)


# --- writing -------------------------------------------------------------------


def _validate_cells(headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Reject non-scalar cells and rows whose width does not match the headers.

    The width check has no schema equivalent (Pydantic's ``TabularCell`` union
    validates one cell in isolation; it cannot see ``headers`` to catch a
    ragged row) and is load-bearing on every path. The per-cell type check
    duplicates what the MCP tool's ``TabularCell`` schema already enforces for
    an MCP-originated call, but this function is also called directly (see the
    unit tests) with values that never passed through that schema — so it
    stays as this leaf's own contract, not dead code.
    """
    for row_index, row in enumerate(rows):
        if len(row) != len(headers):
            raise ValueError(
                f"row {row_index} has {len(row)} cells but there are {len(headers)} headers; "
                f"every row must have exactly one cell per header"
            )
        for column_index, cell in enumerate(row):
            if cell is None or isinstance(cell, str | bool | int):
                continue
            if isinstance(cell, float):
                if math.isfinite(cell):
                    continue
                raise ValueError(
                    f"cell at row {row_index}, column {column_index} is {cell!r}; "
                    f"a cell must be a finite number"
                )
            raise ValueError(
                f"cell at row {row_index}, column {column_index} has unsupported type "
                f"{type(cell).__name__}; a cell must be a string, number, boolean, or null"
            )


def _walk_rows(rows: list[list[TabularCell]], check: Callable[[TabularCell, str], None]) -> None:
    """Call ``check(cell, where)`` for every data cell, in row-major order."""
    for row_index, row in enumerate(rows):
        for column_index, cell in enumerate(row):
            check(cell, f"the cell at row {row_index}, column {column_index}")


def _walk_headers_and_rows(
    headers: list[str],
    rows: list[list[TabularCell]],
    check: Callable[[TabularCell, str], None],
) -> None:
    """Call ``check(cell, where)`` for every header, then every data cell."""
    for index, header in enumerate(headers):
        check(header, f"header {index}")
    _walk_rows(rows, check)


def _reject_csv_formulas(headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Reject strings a spreadsheet would read back as a formula.

    A CSV field carries no type, so a leading ``=``/``+``/``-``/``@`` is
    indistinguishable from a formula when the file is reopened in Excel or
    Calc. Since the alternative would be to alter the value (quoting or
    prefixing it), the write is refused instead. Numbers are unaffected: send
    ``-5`` as the number ``-5``, not the string ``"-5"``.
    """

    def check(value: TabularCell, where: str) -> None:
        if not isinstance(value, str):
            return
        # A leading U+FEFF is not whitespace, so plain .lstrip() leaves it in
        # place; written verbatim by the utf-8 (non -sig) writer below, those
        # bytes ARE a file BOM, which spreadsheet readers consume before
        # parsing what follows — silently turning a would-be-rejected formula
        # into one that reads back as executable.
        stripped = value.lstrip("\ufeff").lstrip()
        if stripped.startswith(CSV_FORMULA_PREFIXES):
            prefixes = "".join(CSV_FORMULA_PREFIXES)
            raise ValueError(
                f"{where} is the string {value!r}, which a spreadsheet would read back as a "
                f"formula (a CSV field cannot say 'this is literal text'). CSV rejects strings "
                f"starting with any of {prefixes!r}. Send a number as a numeric cell rather "
                f"than a string, or write .xlsx, which stores the text literally."
            )

    _walk_headers_and_rows(headers, rows, check)


def _reject_oversized_xlsx_strings(headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Reject strings past Excel's per-cell limit, which openpyxl would silently truncate."""

    def check(value: TabularCell, where: str) -> None:
        if isinstance(value, str) and len(value) > XLSX_MAX_STRING_LENGTH:
            raise ValueError(
                f"{where} is {len(value)} characters long, over the {XLSX_MAX_STRING_LENGTH}-"
                f"character limit for one XLSX cell; shorten it rather than let it be truncated"
            )

    _walk_headers_and_rows(headers, rows, check)


# XML 1.0's Char production admits only #x9 | #xA | #xD | [#x20-#xD7FF] |
# [#xE000-#xFFFD] | [#x10000-#x10FFFF] — so this is every C0 control other
# than tab/LF/CR, plus the surrogate range and the two BMP noncharacters.
# openpyxl's own ILLEGAL_CHARACTERS_RE (used by Cell.check_string, which
# raises IllegalCharacterError) only covers the C0 controls: a character like
# U+FFFF or a lone surrogate sails through that check, the write "succeeds",
# and the writer emits a numeric character reference the XML spec forbids —
# so the file cannot be re-parsed at all. This is the full excluded range, so
# nothing here can silently produce an unreadable workbook.
_XML_ILLEGAL_CHARACTERS_RE = re.compile("[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff\ufffe\uffff]")


def _reject_illegal_xlsx_characters(headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Reject a string with a character XML cannot represent at all."""

    def check(value: TabularCell, where: str) -> None:
        if isinstance(value, str) and _XML_ILLEGAL_CHARACTERS_RE.search(value):
            raise ValueError(
                f"{where} contains a character XML (and therefore XLSX) cannot "
                f"represent; remove it before writing"
            )

    _walk_headers_and_rows(headers, rows, check)


def _reject_xlsx_carriage_returns(headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Reject a string containing a carriage return, which XML normalizes away on read.

    A bare ``\\r`` is XML-legal (unlike the characters
    ``_reject_illegal_xlsx_characters`` rejects), so openpyxl writes it verbatim and the
    write "succeeds" — but XML 1.0's end-of-line handling (Char production, section 2.11)
    requires every conformant parser to normalize a lone CR or a CRLF pair to a single LF
    while parsing, not just this file's own read path. ``"a\\r\\nb"`` and ``"a\\rb"``
    therefore both come back as ``"a\\nb"`` from any XML-compliant reader, openpyxl
    included. Since the alternative would be to silently change the value on write, the
    write is refused instead. CSV has no such normalization (this file's writer/reader pair
    both pass ``newline=""``), so this is XLSX-specific.
    """

    def check(value: TabularCell, where: str) -> None:
        if isinstance(value, str) and "\r" in value:
            raise ValueError(
                f"{where} contains a carriage return, which XML normalizes to a plain "
                f"newline on read (XLSX cannot store \\r or \\r\\n as written); use \\n "
                f"instead, or write .csv, which preserves it exactly"
            )

    _walk_headers_and_rows(headers, rows, check)


def _reject_columnless_xlsx(headers: list[str]) -> None:
    """Reject a zero-column XLSX write: openpyxl drops every row of it on read.

    ``_validate_cells`` already requires every row's width to match
    ``len(headers)``, so a row with zero cells can only occur when ``headers``
    itself is empty — every row, including the header, is then a bare ``<row>``
    element with no ``<c>`` children. openpyxl derives its saved ``<dimension>``
    purely from cells that were actually assigned a value; with none anywhere
    in the sheet, it falls back to a 1x1 ``"A1:A1"``, and the read-only reader
    trusts that declared bound to cap iteration — silently dropping every row
    past the first. CSV has no such bound (a blank line reads back as ``[]``
    just fine), so this is XLSX-specific.
    """
    if not headers:
        raise ValueError(
            "an XLSX write needs at least one column: a zero-column table has no cells "
            "anywhere in the sheet, and openpyxl silently drops every row of it on read"
        )


def _reject_xlsx_empty_strings(rows: list[list[TabularCell]]) -> None:
    """Reject an empty-string row cell, which XLSX cannot tell apart from null.

    openpyxl's writer never emits an inline-string element for a ``""`` cell
    (``etree_write_cell``/``lxml_write_cell`` skip any cell whose value is
    ``None`` or ``""``), so it always reads back as ``None`` regardless of the
    cell's declared type. Since the alternative would be to silently change
    the value on write, the write is refused instead — send ``null`` for "no
    value". Headers are exempt: ``_normalize_header`` already documents that a
    blank header becomes a positional name, so that collapse is intentional,
    not silent loss.
    """

    def check(cell: TabularCell, where: str) -> None:
        if cell == "":
            raise ValueError(
                f"{where} is an empty string, which XLSX cannot distinguish from null "
                f"and always reads back as null; send null instead of an empty string"
            )

    _walk_rows(rows, check)


def _reject_lossy_xlsx_numbers(rows: list[list[TabularCell]]) -> None:
    """Reject a number XLSX's numeric write format cannot hold exactly, or whose
    int/float type would silently flip on read-back.

    openpyxl serializes every numeric cell — int or float alike — through
    ``"%.16g" % value`` (``openpyxl.compat.strings.safe_string``), a fixed
    16-significant-digit text format, with no error when a value needs more
    digits than that to round-trip: an integer past 2**53 or a float needing a
    17th significant digit comes back changed. Detect it by formatting through
    the same ``%.16g`` openpyxl will use and checking the value survives, and
    refuse the write rather than let it through.

    XLSX also has no separate int/float cell type: whether a value reads back
    as ``int`` or ``float`` is inferred purely from whether that same
    ``%.16g`` text contains a ``.``/``e`` — so an integral float like ``1.0``
    formats as ``"1"`` and reads back an ``int``, and a large int like
    ``10**16`` formats as ``"1e+16"`` and reads back a ``float``. There is no
    write-time knob to force the other shape (a cell's number_format is
    cosmetic display only; it does not change the stored ``<v>`` text), so
    this is refused too rather than silently changing the cell's type.
    """

    def check(cell: TabularCell, where: str) -> None:
        if isinstance(cell, bool) or not isinstance(cell, int | float):
            return
        try:
            formatted = f"{cell:.16g}"
        except OverflowError as exc:
            # An int with no float equivalent at all (e.g. 10**400):
            # "%.16g" converts through float internally, which raises
            # OverflowError — not a ValueError — for a magnitude beyond
            # what an IEEE 754 double can hold.
            raise ValueError(
                f"{where} is {cell!r}, too large in magnitude for XLSX's numeric format "
                f"(an IEEE 754 double) to represent at all; send it as a string instead"
            ) from exc
        if float(formatted) != cell:
            raise ValueError(
                f"{where} is {cell!r}, which XLSX's 16-significant-digit numeric format "
                f"cannot represent exactly; send it as a string instead, or reduce its "
                f"precision"
            )
        looks_integral = not any(marker in formatted for marker in ".eE")
        if looks_integral != isinstance(cell, int):
            was = "an int" if isinstance(cell, int) else "a float"
            becomes = "an int" if looks_integral else "a float"
            raise ValueError(
                f"{where} is {cell!r}, {was}, but XLSX would write it as {formatted!r} and "
                f"read it back as {becomes}, silently changing its type; send it as a "
                f"string instead if the type must be preserved exactly"
            )

    _walk_rows(rows, check)


def _write_csv(path: Path, headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Write the fixed comma/quote dialect. ``None`` becomes an empty field."""
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _write_xlsx(path: Path, headers: list[str], rows: list[list[TabularCell]]) -> None:
    """Write one sheet, storing every string as an explicit string cell.

    openpyxl's value setter infers a leading ``=`` as a FORMULA (``data_type``
    ``"f"``). Forcing ``data_type = "s"`` after assignment writes the text as
    an inline string instead, so ``"=1+1"`` round-trips as the literal text it
    was — the server never emits executable spreadsheet code.
    """
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WRITE_SHEET_NAME
    records: list[list[TabularCell]] = [[*headers], *rows]
    for row_index, values in enumerate(records, start=1):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.value = value
            if isinstance(value, str):
                cell.data_type = "s"
    workbook.save(path)


def write_tabular_data(
    headers: list[str],
    rows: list[list[TabularCell]],
    target_path: Path,
    *,
    overwrite: bool = False,
) -> TabularWriteResult:
    """Write ``headers``/``rows`` to a ``.xlsx`` or ``.csv`` file, atomically.

    Everything that could reject the write — path, row widths, cell types,
    formula safety, XLSX string lengths — is checked BEFORE any file is
    created, so a refused write leaves the filesystem untouched.

    The commit is atomic and, with ``overwrite=False``, cannot clobber a target
    that appeared after validation: the file is staged in the target's own
    directory and published with ``os.link``, which fails if the name already
    exists. ``overwrite=True`` commits with ``os.replace`` instead. The staged
    file is removed on every path, and its removal is best-effort: it never
    overrides the outcome (success or a translated error) of the write itself.
    The staging name comes from ``tempfile.mkstemp`` rather than the target
    name, so it stays short regardless of how long the target's own filename
    is.
    """
    resolved = validate_tabular_write_path(target_path)
    fmt = _format_for(resolved)
    _validate_cells(headers, rows)
    if fmt == "csv":
        _reject_csv_formulas(headers, rows)
    else:
        _reject_columnless_xlsx(headers)
        _reject_oversized_xlsx_strings(headers, rows)
        _reject_xlsx_empty_strings(rows)
        _reject_lossy_xlsx_numbers(rows)
        _reject_illegal_xlsx_characters(headers, rows)
        _reject_xlsx_carriage_returns(headers, rows)

    # Diagnostic only — it gives a clear early refusal, but it is NOT the
    # no-clobber gate: the file could still appear between here and the commit.
    # os.link below is the authoritative gate.
    if resolved.exists() and not overwrite:
        raise ValueError(
            f"refusing to overwrite the existing file at {resolved}; "
            f"pass overwrite=true to replace it."
        )

    try:
        fd, staging_name = tempfile.mkstemp(dir=resolved.parent, prefix=".tabular-staging-")
        os.close(fd)
    except OSError as exc:
        # e.g. a permission-denied staging directory. Same translation as the
        # write failure below — nothing has been created yet.
        raise ValueError(f"cannot write {resolved}: {exc}") from exc
    staging = Path(staging_name)
    try:
        try:
            if fmt == "csv":
                _write_csv(staging, headers, rows)
            else:
                _write_xlsx(staging, headers, rows)
        except OSError as exc:
            # e.g. a permission-denied staging directory or a full disk. Not the
            # no-clobber race the os.replace/os.link handling below guards
            # against — this is the write itself failing — but it needs the
            # same translation to reach the tools' ValueError-only boundary
            # instead of escaping as a raw OSError.
            raise ValueError(f"cannot write {resolved}: {exc}") from exc

        try:
            # Hashed here — the staged bytes, before publish — rather than
            # reading ``resolved`` back after the commit: a read failure (e.g.
            # a hostile umask leaving the staged file unreadable) then aborts
            # the whole write with a translated error instead of leaving an
            # already-published target behind that the call reports as failed.
            digest = path_sha256(staging)
        except OSError as exc:
            raise ValueError(f"cannot hash the staged write for {resolved}: {exc}") from exc

        if overwrite:
            try:
                os.replace(staging, resolved)
            except OSError as exc:
                # e.g. a permission-denied target or a filesystem that
                # rejects the replace outright. Translate it the same way as
                # the no-clobber os.link failure below rather than let it
                # escape as a raw OSError past the tools' ValueError-only
                # boundary.
                raise ValueError(f"cannot replace {resolved}: {exc}") from exc
        else:
            try:
                os.link(staging, resolved)
            except FileExistsError as exc:
                # The target was created after the check above. The existing
                # file wins and stays byte-for-byte untouched — never fall back
                # to replacing it.
                raise ValueError(
                    f"refusing to overwrite the existing file at {resolved}; "
                    f"pass overwrite=true to replace it."
                ) from exc
            except OSError as exc:
                # Some filesystems/mounts (e.g. certain network shares) reject
                # hard links entirely (EPERM/ENOTSUP), which is not the
                # no-clobber race this no-overwrite path is meant to guard
                # against. Translate it the same way rather than let it escape
                # as a raw OSError past the tools' ValueError-only boundary.
                raise ValueError(
                    f"cannot publish {resolved}: the filesystem rejected the "
                    f"no-clobber hard-link commit ({exc}); pass overwrite=true "
                    f"to write with a plain replace instead."
                ) from exc
    finally:
        # Best-effort: a failure here (e.g. the same ENAMETOOLONG/EACCES class
        # of error the commit above already guards against) must never replace
        # whatever the try block just raised, nor report a successful publish
        # as a failure.
        try:
            staging.unlink(missing_ok=True)
        except OSError:
            pass

    return TabularWriteResult(
        message=f"Wrote {len(rows)} row(s) and {len(headers)} column(s) to {resolved}.",
        target_path=str(resolved),
        sha256=digest,
        format=fmt,
        rows_written=len(rows),
    )
